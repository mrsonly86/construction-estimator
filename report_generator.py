#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Module tạo báo cáo cho phần mềm dự toán xây dựng
"""

import pandas as pd
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os
from datetime import datetime

class ReportGenerator:
    def __init__(self):
        self.setup_fonts()
    
    def setup_fonts(self):
        """Thiết lập font tiếng Việt"""
        try:
            # Thử load font tiếng Việt (có thể cần tải font riêng)
            # font_path = "fonts/DejaVuSans.ttf"
            # if os.path.exists(font_path):
            #     pdfmetrics.registerFont(TTFont('Vietnamese', font_path))
            pass
        except:
            pass
    
    def export_to_excel(self, project_data, items_data, filename):
        """Xuất dữ liệu ra file Excel"""
        try:
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Sheet thông tin dự án
                project_df = pd.DataFrame([project_data])
                project_df.to_excel(writer, sheet_name='Thông tin dự án', index=False)
                
                # Sheet chi tiết dự toán
                items_df = pd.DataFrame(items_data, columns=[
                    'STT', 'Mã định mức', 'Tên hạng mục', 'Đơn vị', 
                    'Khối lượng', 'Đơn giá', 'Thành tiền', 'Ghi chú'
                ])
                items_df.to_excel(writer, sheet_name='Chi tiết dự toán', index=False)
                
                # Tính toán tổng kết
                total_amount = sum([float(str(item[6]).replace(',', '')) for item in items_data if item[6]])
                vat_amount = total_amount * 0.1
                final_amount = total_amount + vat_amount
                
                summary_data = [
                    ['Tổng số hạng mục', len(items_data)],
                    ['Tổng giá trị (VNĐ)', f"{total_amount:,.0f}"],
                    ['VAT 10% (VNĐ)', f"{vat_amount:,.0f}"],
                    ['Tổng cộng có VAT (VNĐ)', f"{final_amount:,.0f}"]
                ]
                
                summary_df = pd.DataFrame(summary_data, columns=['Mục', 'Giá trị'])
                summary_df.to_excel(writer, sheet_name='Tổng kết', index=False)
                
                # Format Excel
                workbook = writer.book
                
                # Format cho sheet chi tiết
                worksheet = writer.sheets['Chi tiết dự toán']
                
                # Auto-adjust column width
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Định dạng header
                from openpyxl.styles import Font, PatternFill, Alignment
                
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                
                # Định dạng số
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    # Khối lượng (cột 5)
                    if row[4].value:
                        row[4].number_format = '#,##0.00'
                    # Đơn giá (cột 6)
                    if row[5].value:
                        row[5].number_format = '#,##0'
                    # Thành tiền (cột 7)
                    if row[6].value:
                        row[6].number_format = '#,##0'
            
            return True, "Xuất Excel thành công!"
            
        except Exception as e:
            return False, f"Lỗi xuất Excel: {str(e)}"
    
    def export_to_pdf(self, project_data, items_data, filename):
        """Xuất dữ liệu ra file PDF"""
        try:
            doc = SimpleDocTemplate(filename, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Title'],
                fontSize=18,
                spaceAfter=30,
                alignment=1,  # Center
                textColor=colors.darkblue
            )
            
            title = Paragraph("BÁO CÁO DỰ TOÁN XÂY DỰNG", title_style)
            story.append(title)
            story.append(Spacer(1, 20))
            
            # Project info
            project_info = f"""
            <b>Tên dự án:</b> {project_data.get('name', '')}<br/>
            <b>Địa điểm:</b> {project_data.get('location', '')}<br/>
            <b>Chủ đầu tư:</b> {project_data.get('client', '')}<br/>
            <b>Ngày tạo:</b> {project_data.get('date', '')}<br/>
            """
            
            info_paragraph = Paragraph(project_info, styles['Normal'])
            story.append(info_paragraph)
            story.append(Spacer(1, 20))
            
            # Table data
            table_data = [['STT', 'Mã định mức', 'Tên hạng mục', 'Đơn vị', 'Khối lượng', 'Đơn giá', 'Thành tiền']]
            
            for item in items_data:
                row = [
                    str(item[0]), str(item[1]), str(item[2]), str(item[3]),
                    str(item[4]), str(item[5]), str(item[6])
                ]
                table_data.append(row)
            
            # Create table
            table = Table(table_data, colWidths=[0.5*inch, 1*inch, 2.5*inch, 0.7*inch, 0.8*inch, 1*inch, 1*inch])
            
            # Table style
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
            story.append(Spacer(1, 20))
            
            # Summary
            total_amount = sum([float(str(item[6]).replace(',', '')) for item in items_data if item[6]])
            vat_amount = total_amount * 0.1
            final_amount = total_amount + vat_amount
            
            summary_text = f"""
            <b>TỔNG KẾT:</b><br/>
            Tổng số hạng mục: {len(items_data)}<br/>
            Tổng giá trị: {total_amount:,.0f} VNĐ<br/>
            VAT (10%): {vat_amount:,.0f} VNĐ<br/>
            <b>Tổng cộng (có VAT): {final_amount:,.0f} VNĐ</b>
            """
            
            summary_paragraph = Paragraph(summary_text, styles['Normal'])
            story.append(summary_paragraph)
            
            # Build PDF
            doc.build(story)
            
            return True, "Xuất PDF thành công!"
            
        except Exception as e:
            return False, f"Lỗi xuất PDF: {str(e)}"
    
    def create_detailed_report(self, project_data, items_data, norms_data, filename):
        """Tạo báo cáo chi tiết với phân tích"""
        try:
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Sheet 1: Thông tin dự án
                project_df = pd.DataFrame([project_data])
                project_df.to_excel(writer, sheet_name='Thông tin dự án', index=False)
                
                # Sheet 2: Chi tiết dự toán
                items_df = pd.DataFrame(items_data, columns=[
                    'STT', 'Mã định mức', 'Tên hạng mục', 'Đơn vị', 
                    'Khối lượng', 'Đơn giá', 'Thành tiền', 'Ghi chú'
                ])
                items_df.to_excel(writer, sheet_name='Chi tiết dự toán', index=False)
                
                # Sheet 3: Phân tích theo loại công việc
                category_analysis = {}
                for item in items_data:
                    norm_code = item[1]
                    amount = float(str(item[6]).replace(',', '')) if item[6] else 0
                    
                    # Tìm category từ norm_code
                    category = self.get_category_from_code(norm_code)
                    if category not in category_analysis:
                        category_analysis[category] = {'count': 0, 'total': 0}
                    
                    category_analysis[category]['count'] += 1
                    category_analysis[category]['total'] += amount
                
                analysis_data = []
                for category, data in category_analysis.items():
                    analysis_data.append([category, data['count'], f"{data['total']:,.0f}"])
                
                analysis_df = pd.DataFrame(analysis_data, columns=['Loại công việc', 'Số hạng mục', 'Tổng giá trị (VNĐ)'])
                analysis_df.to_excel(writer, sheet_name='Phân tích theo loại', index=False)
                
                # Sheet 4: Định mức sử dụng
                norms_df = pd.DataFrame(norms_data, columns=[
                    'Mã', 'Tên định mức', 'Đơn vị', 'Vật liệu', 'Nhân công', 'Máy móc', 'Tổng cộng', 'Loại'
                ])
                norms_df.to_excel(writer, sheet_name='Định mức sử dụng', index=False)
                
                # Sheet 5: Tổng kết
                total_amount = sum([float(str(item[6]).replace(',', '')) for item in items_data if item[6]])
                vat_amount = total_amount * 0.1
                final_amount = total_amount + vat_amount
                
                summary_data = [
                    ['Tổng số hạng mục', len(items_data), ''],
                    ['Tổng giá trị (VNĐ)', f"{total_amount:,.0f}", ''],
                    ['VAT 10% (VNĐ)', f"{vat_amount:,.0f}", ''],
                    ['Tổng cộng có VAT (VNĐ)', f"{final_amount:,.0f}", 'Tổng cuối cùng'],
                    ['', '', ''],
                    ['Ngày tạo báo cáo', datetime.now().strftime("%d/%m/%Y %H:%M"), ''],
                    ['Phần mềm', 'Dự toán Xây dựng Chuyên nghiệp v1.0', 'Miễn phí']
                ]
                
                summary_df = pd.DataFrame(summary_data, columns=['Mục', 'Giá trị', 'Ghi chú'])
                summary_df.to_excel(writer, sheet_name='Tổng kết', index=False)
            
            return True, "Tạo báo cáo chi tiết thành công!"
            
        except Exception as e:
            return False, f"Lỗi tạo báo cáo: {str(e)}"
    
    def get_category_from_code(self, norm_code):
        """Xác định loại công việc từ mã định mức"""
        if not norm_code:
            return "Khác"
        
        code_upper = norm_code.upper()
        
        if code_upper.startswith('A.1'):
            return "Đào đất"
        elif code_upper.startswith('A.2'):
            return "Bê tông"
        elif code_upper.startswith('A.3'):
            return "Thép"
        elif code_upper.startswith('A.4'):
            return "Gạch xây"
        elif code_upper.startswith('A.5'):
            return "Hoàn thiện"
        elif code_upper.startswith('B.'):
            return "Kết cấu"
        elif code_upper.startswith('C.'):
            return "Hệ thống M&E"
        else:
            return "Khác"